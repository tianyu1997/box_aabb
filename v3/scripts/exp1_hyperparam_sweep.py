#!/usr/bin/env python3
"""
scripts/exp1_hyperparam_sweep.py — SBF-Dijkstra 超参数扫描

实验1: 在 2dof 与 panda 场景下 (各 5 个随机障碍物) 搜索 SBF-Dijkstra
的稳定高性能参数。

扫描参数:
  - ffb_min_edge
  - max_boxes
  - max_consecutive_miss  (即 max_failed_samples)
  - guided_sample_ratio

记录指标:
  - 成功率
  - 中位/P90 规划时间
  - 平均 box 数
  - 平均路径长度

输出:
  - Excel 含 summary + 2dof + panda 三个 Sheet

用法:
    cd v3
    python scripts/exp1_hyperparam_sweep.py --seeds 30 --start-seed 0
    python scripts/exp1_hyperparam_sweep.py --seeds 1 --start-seed 42   # trial
"""
from __future__ import annotations

import argparse
import itertools
import json
import math
import sys
import time
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Tuple

import numpy as np

# Windows UTF-8
if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")

_ROOT = Path(__file__).resolve().parents[1]  # v3/
_SRC = _ROOT / "src"
for p in (_ROOT, _SRC):
    sp = str(p)
    if sp not in sys.path:
        sys.path.insert(0, sp)

from aabb.robot import load_robot
from forest.scene import Scene
from forest.collision import CollisionChecker
from baselines import SBFAdapter
from planner.pipeline import PandaGCSConfig, build_panda_scene

# ═══════════════════════════════════════════════════════════════════
# Helpers
# ═══════════════════════════════════════════════════════════════════

def _compute_period(joint_limits):
    spans = [hi - lo for lo, hi in joint_limits]
    if not spans:
        return None
    span0 = spans[0]
    all_same = all(abs(s - span0) < 1e-6 for s in spans)
    return float(span0) if (all_same and abs(span0 - 2 * math.pi) < 0.1) else None


def _path_length(waypoints, period=None):
    if waypoints is None or len(waypoints) < 2:
        return float('nan')
    if period is not None:
        half = period / 2.0
        total = 0.0
        for i in range(len(waypoints) - 1):
            diff = ((waypoints[i + 1] - waypoints[i]) + half) % period - half
            total += float(np.linalg.norm(diff))
        return total
    total = 0.0
    for i in range(len(waypoints) - 1):
        total += float(np.linalg.norm(waypoints[i + 1] - waypoints[i]))
    return total


# ═══════════════════════════════════════════════════════════════════
# Scene builders
# ═══════════════════════════════════════════════════════════════════

def build_random_2d_scene(robot, q_start, q_goal, rng, n_obstacles=5,
                          max_trials=500):
    """Build a random 2D scene with n obstacles, ensuring feasibility."""
    from baselines.rrt_family import plan_rrt_connect

    jl = robot.joint_limits
    period = _compute_period(jl)

    for trial_i in range(max_trials):
        scene = Scene()
        for i in range(n_obstacles):
            cx = float(rng.uniform(-1.8, 1.8))
            cy = float(rng.uniform(-1.8, 1.8))
            w = float(rng.uniform(0.3, 0.8))
            h = float(rng.uniform(0.3, 0.8))
            scene.add_obstacle([cx - w / 2, cy - h / 2],
                               [cx + w / 2, cy + h / 2], name=f"obs_{i}")
        checker = CollisionChecker(robot=robot, scene=scene)
        if checker.check_config_collision(q_start):
            continue
        if checker.check_config_collision(q_goal):
            continue
        # reject trivially solvable (direct path free)
        if not checker.check_segment_collision(q_start, q_goal, 0.03):
            continue
        if period is not None and not checker.check_segment_collision(
                q_start, q_goal, 0.03, period=period):
            continue
        # feasibility check
        res = plan_rrt_connect(
            q_start, q_goal, jl, checker,
            timeout=2.0, step_size=0.3, resolution=0.05,
            seed=trial_i + 7777, period=period)
        if not res['success']:
            continue
        return scene
    raise RuntimeError(f"Cannot build 2D scene after {max_trials} trials")


def build_panda_5obs_scene(robot, q_start, q_goal, scene_seed):
    """Build a Panda scene with 5 random obstacles, ensuring feasibility."""
    cfg = PandaGCSConfig()
    cfg.n_obstacles = 5
    cfg.workspace_radius = 0.85
    cfg.workspace_z_range = (0.0, 1.0)
    cfg.obs_size_range = (0.08, 0.25)
    rng = np.random.default_rng(scene_seed)
    scene = build_panda_scene(rng, cfg, robot=robot,
                              q_start=q_start, q_goal=q_goal)
    return scene


# ═══════════════════════════════════════════════════════════════════
# Sweep grid
# ═══════════════════════════════════════════════════════════════════

SWEEP_PARAMS_2DOF = {
    'ffb_min_edge':        [0.1, 0.05, 0.15, 0.01],
    'max_boxes':           [99999],              # fixed; max_consecutive_miss controls exit
    'max_consecutive_miss': [3, 5, 10],
    'guided_sample_ratio': [0.8, 0.6, 0.9],
}

SWEEP_PARAMS_PANDA = {
    'ffb_min_edge':        [ 0.05,0.075, 0.10, 0.125],
    'max_boxes':           [1500],
    'max_consecutive_miss': [9999],               # fixed; max_boxes controls exit
    'guided_sample_ratio': [0.3, 0.4, 0.2],
}

# unified key list for Excel (superset)
SWEEP_PARAMS = SWEEP_PARAMS_PANDA  # default reference for param_keys


def generate_param_combos(sweep: dict) -> List[dict]:
    """Full factorial sweep."""
    keys = list(sweep.keys())
    vals = list(sweep.values())
    combos = []
    for combo in itertools.product(*vals):
        combos.append(dict(zip(keys, combo)))
    return combos


# ═══════════════════════════════════════════════════════════════════
# Per-combo evaluation
# ═══════════════════════════════════════════════════════════════════

def evaluate_combo(robot, scenes: List[Tuple], params: dict,
                   timeout: float = 30.0) -> dict:
    """Run SBF-Dijkstra on all (scene, q_start, q_goal) with given params.

    Args:
        scenes: List of (scene, q_start, q_goal, seed) tuples
        params: Override params for SBF config
    Returns:
        dict with aggregated metrics
    """
    results = []
    for scene, q_start, q_goal, seed in scenes:
        sbf = SBFAdapter(method="dijkstra")
        cfg = dict(params)
        cfg['seed'] = seed
        sbf.setup(robot, scene, cfg)
        try:
            t0 = time.perf_counter()
            res = sbf.plan(q_start, q_goal, timeout=timeout)
            dt = time.perf_counter() - t0
        except Exception as e:
            results.append({
                'success': False, 'time_ms': 0, 'path_length': float('nan'),
                'n_boxes': 0, 'seed': seed, 'error': str(e),
            })
            continue

        period = _compute_period(robot.joint_limits)
        pl = _path_length(res.path, period) if res.success else float('nan')
        n_boxes = res.nodes_explored

        results.append({
            'success': res.success,
            'time_ms': dt * 1000,
            'path_length': pl,
            'n_boxes': n_boxes,
            'seed': seed,
        })

    # aggregate
    n_total = len(results)
    successes = [r for r in results if r['success']]
    n_success = len(successes)

    times = sorted([r['time_ms'] for r in successes]) if successes else []
    path_lengths = [r['path_length'] for r in successes
                    if not math.isnan(r['path_length'])]
    box_counts = [r['n_boxes'] for r in successes]

    return {
        'params': params,
        'n_total': n_total,
        'n_success': n_success,
        'success_rate': n_success / n_total if n_total > 0 else 0,
        'median_time_ms': float(np.median(times)) if times else float('nan'),
        'p90_time_ms': float(np.percentile(times, 90)) if times else float('nan'),
        'mean_time_ms': float(np.mean(times)) if times else float('nan'),
        'mean_path_length': float(np.mean(path_lengths)) if path_lengths else float('nan'),
        'mean_n_boxes': float(np.mean(box_counts)) if box_counts else float('nan'),
        'per_seed': results,
    }


# ═══════════════════════════════════════════════════════════════════
# Excel output
# ═══════════════════════════════════════════════════════════════════

def write_excel(results_2d: List[dict], results_panda: List[dict],
                out_path: Path):
    """Write sweep results to Excel with summary + per-env sheets."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ── Style helpers ──
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill(start_color="4472C4", end_color="4472C4",
                           fill_type="solid")
    hdr_align = Alignment(horizontal="center", vertical="center",
                          wrap_text=True)
    even_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3",
                            fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="B4C6E7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    good_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE",
                            fill_type="solid")
    best_font = Font(bold=True, color="006100")

    def _style_header(ws, row, ncols):
        for col in range(1, ncols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = hdr_align

    def _style_data(ws, start_row, end_row, ncols):
        for r in range(start_row, end_row + 1):
            for c in range(1, ncols + 1):
                cell = ws.cell(row=r, column=c)
                cell.alignment = center_align
                cell.border = border
                if (r - start_row) % 2 == 1:
                    cell.fill = even_fill

    param_keys = list(SWEEP_PARAMS.keys())

    def _write_env_sheet(ws, env_results: List[dict], env_name: str):
        """Write one environment's results to a sheet."""
        headers = (
            param_keys +
            ["Success Rate", "Median Time\n(ms)", "P90 Time\n(ms)",
             "Mean Time\n(ms)", "Mean Path\nLength", "Mean\nBoxes"]
        )
        ncols = len(headers)
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h)
        _style_header(ws, 1, ncols)

        # Sort: success_rate desc, median_time asc
        sorted_results = sorted(env_results,
                                key=lambda x: (-x['success_rate'],
                                               x['median_time_ms']
                                               if not math.isnan(x['median_time_ms'])
                                               else 1e9))

        # Find best combo (highest success, lowest median time)
        best_idx = 0 if sorted_results else -1

        row = 2
        for i, res in enumerate(sorted_results):
            for j, pk in enumerate(param_keys):
                ws.cell(row=row, column=j + 1, value=res['params'][pk])
            col_off = len(param_keys)
            ws.cell(row=row, column=col_off + 1,
                    value=f"{res['success_rate']:.0%}")
            ws.cell(row=row, column=col_off + 2,
                    value=round(res['median_time_ms'], 1)
                    if not math.isnan(res['median_time_ms']) else "N/A")
            ws.cell(row=row, column=col_off + 3,
                    value=round(res['p90_time_ms'], 1)
                    if not math.isnan(res['p90_time_ms']) else "N/A")
            ws.cell(row=row, column=col_off + 4,
                    value=round(res['mean_time_ms'], 1)
                    if not math.isnan(res['mean_time_ms']) else "N/A")
            ws.cell(row=row, column=col_off + 5,
                    value=round(res['mean_path_length'], 3)
                    if not math.isnan(res['mean_path_length']) else "N/A")
            ws.cell(row=row, column=col_off + 6,
                    value=round(res['mean_n_boxes'], 0)
                    if not math.isnan(res['mean_n_boxes']) else "N/A")

            # Highlight best
            if i == best_idx:
                for c in range(1, ncols + 1):
                    ws.cell(row=row, column=c).fill = good_fill
                    ws.cell(row=row, column=c).font = best_font

            row += 1

        _style_data(ws, 2, row - 1, ncols)

        # Column widths
        for i in range(1, ncols + 1):
            ws.column_dimensions[get_column_letter(i)].width = 18
        ws.freeze_panes = "A2"

    # ── Summary sheet ──
    ws_sum = wb.active
    ws_sum.title = "summary"
    sum_headers = ["Environment"] + param_keys + [
        "Success Rate", "Median Time (ms)", "P90 Time (ms)",
        "Mean Path Length", "Mean Boxes"
    ]
    ncols_sum = len(sum_headers)
    for col, h in enumerate(sum_headers, 1):
        ws_sum.cell(row=1, column=col, value=h)
    _style_header(ws_sum, 1, ncols_sum)

    row_sum = 2
    for env_name, env_results in [("2dof", results_2d),
                                   ("panda", results_panda)]:
        if not env_results:
            continue
        # Find best combo for this env
        best = min(env_results,
                   key=lambda x: (-x['success_rate'],
                                  x['median_time_ms']
                                  if not math.isnan(x['median_time_ms'])
                                  else 1e9))
        ws_sum.cell(row=row_sum, column=1, value=env_name)
        for j, pk in enumerate(param_keys):
            ws_sum.cell(row=row_sum, column=2 + j, value=best['params'][pk])
        off = 2 + len(param_keys)
        ws_sum.cell(row=row_sum, column=off,
                    value=f"{best['success_rate']:.0%}")
        ws_sum.cell(row=row_sum, column=off + 1,
                    value=round(best['median_time_ms'], 1)
                    if not math.isnan(best['median_time_ms']) else "N/A")
        ws_sum.cell(row=row_sum, column=off + 2,
                    value=round(best['p90_time_ms'], 1)
                    if not math.isnan(best['p90_time_ms']) else "N/A")
        ws_sum.cell(row=row_sum, column=off + 3,
                    value=round(best['mean_path_length'], 3)
                    if not math.isnan(best['mean_path_length']) else "N/A")
        ws_sum.cell(row=row_sum, column=off + 4,
                    value=round(best['mean_n_boxes'], 0)
                    if not math.isnan(best['mean_n_boxes']) else "N/A")
        row_sum += 1

    _style_data(ws_sum, 2, row_sum - 1, ncols_sum)
    for i in range(1, ncols_sum + 1):
        ws_sum.column_dimensions[get_column_letter(i)].width = 18
    ws_sum.freeze_panes = "B2"

    # ── Per-env sheets ──
    ws_2d = wb.create_sheet("2dof")
    _write_env_sheet(ws_2d, results_2d, "2dof")

    ws_panda = wb.create_sheet("panda")
    _write_env_sheet(ws_panda, results_panda, "panda")

    wb.save(out_path)
    print(f"\nExcel saved: {out_path}")


# ═══════════════════════════════════════════════════════════════════
# Main
# ═══════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        description="Exp1: SBF-Dijkstra hyperparameter sweep")
    parser.add_argument("--seeds", type=int, default=30,
                        help="Number of random seeds per combo (default: 30)")
    parser.add_argument("--start-seed", type=int, default=0,
                        help="Starting seed (default: 0)")
    parser.add_argument("--timeout", type=float, default=30.0,
                        help="SBF planning timeout (s)")
    parser.add_argument("--scene-seed", type=int, default=1000,
                        help="Fixed seed for scene generation (default: 1000)")
    parser.add_argument("--skip-2dof", action="store_true",
                        help="Skip 2DOF experiments")
    parser.add_argument("--skip-panda", action="store_true",
                        help="Skip Panda experiments")
    args = parser.parse_args()

    seeds = list(range(args.start_seed, args.start_seed + args.seeds))
    combos_2d = generate_param_combos(SWEEP_PARAMS_2DOF)
    combos_panda = generate_param_combos(SWEEP_PARAMS_PANDA)
    print(f"Exp1: Hyperparameter sweep")
    print(f"  Seeds: {args.seeds} ({args.start_seed}..{args.start_seed + args.seeds - 1})")
    print(f"  2DOF combos: {len(combos_2d)}")
    print(f"  Panda combos: {len(combos_panda)}")
    print(f"  Sweep params: {list(SWEEP_PARAMS.keys())}")

    # ── 2DOF ──
    results_2d = []
    if not args.skip_2dof:
        print(f"\n{'='*60}")
        print(f"  2DOF (5 obstacles)")
        print(f"{'='*60}")

        robot_2d = load_robot("2dof_planar")
        q_start_2d = np.array([-2.0, 1.5])
        q_goal_2d  = np.array([2.0, -1.5])

        # Pre-build scenes (one per seed, fixed scene_seed)
        print("  Building scenes ...")
        scenes_2d = []
        for s in seeds:
            # Scene uses scene_seed + planner_seed to keep obstacles fixed
            # per sweep but query seed varies
            rng = np.random.default_rng(args.scene_seed + s)
            try:
                scene = build_random_2d_scene(robot_2d, q_start_2d, q_goal_2d,
                                              rng, n_obstacles=5)
                scenes_2d.append((scene, q_start_2d, q_goal_2d, s))
            except RuntimeError:
                print(f"    seed={s}: SKIP (no valid scene)")

        print(f"  Valid scenes: {len(scenes_2d)}/{len(seeds)}")

        n_combos_2d = len(combos_2d)
        for ci, combo in enumerate(combos_2d, 1):
            print(f"\n  [{ci}/{n_combos_2d}] {combo}")
            res = evaluate_combo(robot_2d, scenes_2d, combo,
                                 timeout=args.timeout)
            print(f"    -> success={res['success_rate']:.0%}  "
                  f"median={res['median_time_ms']:.0f}ms  "
                  f"P90={res['p90_time_ms']:.0f}ms  "
                  f"path={res['mean_path_length']:.3f}  "
                  f"boxes={res['mean_n_boxes']:.0f}")
            results_2d.append(res)

    # ── Panda ──
    results_panda = []
    if not args.skip_panda:
        print(f"\n{'='*60}")
        print(f"  Panda 7-DOF (5 obstacles)")
        print(f"{'='*60}")

        robot_panda = load_robot("panda")
        q_start_panda = np.array([0.5, -1.2, 0.5, -2.5, 0.5, 0.8, 1.5])
        q_goal_panda  = np.array([-2.0, 1.2, -1.8, -0.5, -2.0, 3.0, -1.8])

        print("  Building scenes ...")
        scenes_panda = []
        for s in seeds:
            try:
                scene = build_panda_5obs_scene(
                    robot_panda, q_start_panda, q_goal_panda,
                    scene_seed=args.scene_seed + s)
                scenes_panda.append((scene, q_start_panda, q_goal_panda, s))
            except RuntimeError:
                print(f"    seed={s}: SKIP (no valid scene)")

        print(f"  Valid scenes: {len(scenes_panda)}/{len(seeds)}")

        n_combos_panda = len(combos_panda)
        for ci, combo in enumerate(combos_panda, 1):
            print(f"\n  [{ci}/{n_combos_panda}] {combo}")
            res = evaluate_combo(robot_panda, scenes_panda, combo,
                                 timeout=args.timeout)
            print(f"    -> success={res['success_rate']:.0%}  "
                  f"median={res['median_time_ms']:.0f}ms  "
                  f"P90={res['p90_time_ms']:.0f}ms  "
                  f"path={res['mean_path_length']:.3f}  "
                  f"boxes={res['mean_n_boxes']:.0f}")
            results_panda.append(res)

    # ── Save JSON raw results ──
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_dir = _ROOT / "experiments" / "output"
    out_dir.mkdir(parents=True, exist_ok=True)

    raw_data = {
        'timestamp': ts,
        'seeds': seeds,
        'sweep_params': SWEEP_PARAMS,
        '2dof': [
            {**r, 'per_seed': [
                {k: v for k, v in ps.items() if k != 'error'}
                for ps in r['per_seed']
            ]}
            for r in results_2d
        ],
        'panda': [
            {**r, 'per_seed': [
                {k: v for k, v in ps.items() if k != 'error'}
                for ps in r['per_seed']
            ]}
            for r in results_panda
        ],
    }
    json_path = out_dir / f"exp1_sweep_{ts}.json"
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(raw_data, f, indent=2, default=str)
    print(f"\nJSON saved: {json_path}")

    # ── Excel ──
    xlsx_path = out_dir / f"exp1_hyperparam_sweep_{ts}.xlsx"
    write_excel(results_2d, results_panda, xlsx_path)

    # ── Console summary ──
    print(f"\n{'='*60}")
    print("  RECOMMENDED PARAMETERS")
    print(f"{'='*60}")
    for env_name, env_results in [("2dof", results_2d),
                                   ("panda", results_panda)]:
        if not env_results:
            continue
        best = min(env_results,
                   key=lambda x: (-x['success_rate'],
                                  x['median_time_ms']
                                  if not math.isnan(x['median_time_ms'])
                                  else 1e9))
        print(f"\n  {env_name}:")
        print(f"    params:       {best['params']}")
        print(f"    success_rate: {best['success_rate']:.0%}")
        print(f"    median_time:  {best['median_time_ms']:.0f} ms")
        print(f"    P90_time:     {best['p90_time_ms']:.0f} ms")
        print(f"    mean_path:    {best['mean_path_length']:.3f}")
        print(f"    mean_boxes:   {best['mean_n_boxes']:.0f}")


if __name__ == "__main__":
    main()
