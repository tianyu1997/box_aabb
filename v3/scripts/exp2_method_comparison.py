#!/usr/bin/env python3
"""
scripts/exp2_method_comparison.py — SBF-Dijkstra vs OMPL 方法对比实验

实验2:
  对比 sbf-dijkstra / ompl-rrt-connect / ompl-bit*(2s) 在 2dof 与 panda
  (各 5 个随机障碍物) 下的表现。

记录:
  1. 各方法纯算法耗时
  2. SBF 细分:
     - 冷启动 (no_cache) 建图耗时
     - 热启动 (有 cache) 建图耗时
     - 新增 1 个障碍后的增量更新耗时
     - 仅修改起终点的重规划耗时 (复用 forest)
     - goal/target tree 连接时 box 数
     - 各子模块耗时明细
  3. BIT* 路径质量优于 SBF 时的耗时

输出:
  Excel: summary / 2dof / panda / sbf_breakdown / abnormal_cases

用法:
    cd v3
    python scripts/exp2_method_comparison.py --seeds 30 --start-seed 0
    python scripts/exp2_method_comparison.py --seeds 1 --start-seed 42  # trial
"""
from __future__ import annotations

import argparse
import json
import math
import subprocess
import sys
import time
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import numpy as np

# Windows UTF-8
if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")

_ROOT = Path(__file__).resolve().parents[1]  # v3/
_SRC = _ROOT / "src"
for p in (_ROOT, _SRC):
    sp = str(p)
    if sp not in sys.path:
        sys.path.insert(0, sp)

from aabb.robot import load_robot
from forest.scene import Scene
from forest.collision import CollisionChecker
from baselines import SBFAdapter
from baselines.ompl_adapter import OMPLPlanner
from planner.pipeline import PandaGCSConfig, build_panda_scene, set_verbose

# 关闭 pipeline 内部的控制台输出
set_verbose(False)

# 2DOF OMPL bridge (2D-specific)
_BRIDGE_2DOF = "/mnt/c/Users/TIAN/Documents/box_aabb/v3/scripts/ompl_bridge_2dof.py"


# ═══════════════════════════════════════════════════════════════════
# Helpers
# ═══════════════════════════════════════════════════════════════════

def _compute_period(joint_limits):
    spans = [hi - lo for lo, hi in joint_limits]
    if not spans:
        return None
    span0 = spans[0]
    all_same = all(abs(s - span0) < 1e-6 for s in spans)
    return float(span0) if (all_same and abs(span0 - 2 * math.pi) < 0.1) else None


def _path_length(waypoints, period=None):
    if waypoints is None or len(waypoints) < 2:
        return float('nan')
    if period is not None:
        half = period / 2.0
        total = 0.0
        for i in range(len(waypoints) - 1):
            diff = ((waypoints[i + 1] - waypoints[i]) + half) % period - half
            total += float(np.linalg.norm(diff))
        return total
    total = 0.0
    for i in range(len(waypoints) - 1):
        total += float(np.linalg.norm(waypoints[i + 1] - waypoints[i]))
    return total


def _find_time_to_reach_cost(cost_history, target_cost,
                              final_cost=None, total_time=None):
    """From BIT* cost_history [(t,cost),...] find first time cost<=target."""
    if not cost_history:
        if final_cost is not None and total_time is not None and final_cost <= target_cost:
            return total_time
        return None
    for t, c in cost_history:
        if c <= target_cost:
            return t
    if final_cost is not None and total_time is not None and final_cost <= target_cost:
        return total_time
    return None


def _safe_float(v, default=float('nan')):
    if v is None:
        return default
    try:
        f = float(v)
        return f if not math.isnan(f) else default
    except (TypeError, ValueError):
        return default


# ═══════════════════════════════════════════════════════════════════
# Scene builders
# ═══════════════════════════════════════════════════════════════════

def build_random_2d_scene(robot, q_start, q_goal, rng, n_obstacles=5,
                          max_trials=500):
    """Build a random 2D scene, ensuring non-trivial but feasible."""
    from baselines.rrt_family import plan_rrt_connect

    jl = robot.joint_limits
    period = _compute_period(jl)

    for trial_i in range(max_trials):
        scene = Scene()
        for i in range(n_obstacles):
            cx = float(rng.uniform(-1.8, 1.8))
            cy = float(rng.uniform(-1.8, 1.8))
            w = float(rng.uniform(0.3, 0.8))
            h = float(rng.uniform(0.3, 0.8))
            scene.add_obstacle([cx - w / 2, cy - h / 2],
                               [cx + w / 2, cy + h / 2], name=f"obs_{i}")
        checker = CollisionChecker(robot=robot, scene=scene)
        if checker.check_config_collision(q_start):
            continue
        if checker.check_config_collision(q_goal):
            continue
        if not checker.check_segment_collision(q_start, q_goal, 0.03):
            continue
        if period is not None and not checker.check_segment_collision(
                q_start, q_goal, 0.03, period=period):
            continue
        res = plan_rrt_connect(
            q_start, q_goal, jl, checker,
            timeout=2.0, step_size=0.3, resolution=0.05,
            seed=trial_i + 7777, period=period)
        if not res['success']:
            continue
        return scene
    raise RuntimeError(f"Cannot build 2D scene after {max_trials} trials")


def build_panda_5obs_scene(robot, q_start, q_goal, scene_seed):
    """Build panda 5-obs scene."""
    cfg = PandaGCSConfig()
    cfg.n_obstacles = 5
    cfg.workspace_radius = 0.85
    cfg.workspace_z_range = (0.0, 1.0)
    cfg.obs_size_range = (0.08, 0.25)
    rng = np.random.default_rng(scene_seed)
    return build_panda_scene(rng, cfg, robot=robot,
                             q_start=q_start, q_goal=q_goal)


# ═══════════════════════════════════════════════════════════════════
# OMPL 2DOF via WSL
# ═══════════════════════════════════════════════════════════════════

def run_ompl_2dof(q_start, q_goal, joint_limits, scene, algorithm,
                  timeout=2.0, seed=42):
    """Call OMPL bridge for 2DOF. Returns PlanningResult-like dict or None."""
    obs_list = []
    for obs in scene.get_obstacles():
        obs_list.append({
            "min_point": obs.min_point.tolist(),
            "max_point": obs.max_point.tolist(),
            "name": obs.name,
        })
    problem = {
        "robot_name": "2dof_planar",
        "q_start": q_start.tolist(),
        "q_goal": q_goal.tolist(),
        "joint_limits": [[float(lo), float(hi)] for lo, hi in joint_limits],
        "obstacles": obs_list,
        "algorithms": [algorithm],
        "timeout": timeout,
        "trials": 1,
        "seed": seed,
        "step_size": 0.3,
    }
    try:
        proc = subprocess.run(
            ["wsl", "python3", _BRIDGE_2DOF],
            input=json.dumps(problem),
            capture_output=True, text=True,
            encoding='utf-8', errors='replace',
            timeout=timeout + 60,
        )
    except (subprocess.TimeoutExpired, FileNotFoundError):
        return None

    stdout = proc.stdout.strip()
    if not stdout:
        return None
    json_start = stdout.find('{')
    if json_start < 0:
        return None
    try:
        raw = json.loads(stdout[json_start:])
    except json.JSONDecodeError:
        return None

    data = raw.get(algorithm, {})
    if "error" in data or not data.get("trials"):
        return None

    t = data["trials"][0]
    if not t.get("success", False):
        return {'success': False, 'algorithm': algorithm}

    wps_raw = data.get("best_waypoints", [])
    return {
        'success': True,
        'algorithm': algorithm,
        'path': [np.array(w) for w in wps_raw] if wps_raw else None,
        'cost': t.get("path_length", float('inf')),
        'planning_time': t.get("plan_time_s", 0),
        'first_solution_time': t.get("first_solution_time", 0),
        'first_solution_cost': t.get("first_solution_cost"),
        'cost_history': t.get("cost_history", []),
        'n_nodes': t.get("n_nodes", 0),
        'n_collision_checks': t.get("n_collision_checks", 0),
    }


# ═══════════════════════════════════════════════════════════════════
# Run SBF with detailed breakdown
# ═══════════════════════════════════════════════════════════════════

# ── Best configs from Exp1 sweep ────────────────────────────
SBF_CFG_2DOF = {
    'ffb_min_edge':        0.1,
    'max_boxes':           99999,   # miss controls exit
    'max_consecutive_miss': 10,
    'guided_sample_ratio': 0.6,
    'n_edge_samples':      3,
}

SBF_CFG_PANDA = {
    'ffb_min_edge':        0.1,
    'max_boxes':           1500,    # max_boxes controls exit
    'max_consecutive_miss': 9999,
    'guided_sample_ratio': 0.3,
    'n_edge_samples':      3,
}


def run_sbf_detailed(robot, scene, q_start, q_goal, seed,
                     q_start_alt, q_goal_alt, timeout=30.0,
                     sbf_cfg_override: dict | None = None):
    """Run SBF-Dijkstra with full breakdown.

    Args:
        sbf_cfg_override: environment-specific SBF hyperparameters.
            If None, uses SBF_CFG_PANDA as default.

    Returns dict with:
      - cold_start: no_cache build+plan result
      - warm_start: with cache build+plan result
      - incremental: add 1 obstacle, re-plan
      - replan: different query on same forest
      - breakdown: sub-phase timing
    """
    period = _compute_period(robot.joint_limits)
    result = {}

    base_cfg = dict(sbf_cfg_override or SBF_CFG_PANDA)
    base_cfg['seed'] = seed

    # ── 1) Cold start (no_cache) ──
    print("    [SBF] cold start (no_cache) ...")
    sbf_cold = SBFAdapter(method="dijkstra")
    cfg_cold = dict(base_cfg)
    cfg_cold['no_cache'] = True
    sbf_cold.setup(robot, scene, cfg_cold)
    try:
        t0 = time.perf_counter()
        res_cold = sbf_cold.plan(q_start, q_goal, timeout=timeout)
        dt_cold = time.perf_counter() - t0
    except Exception as e:
        res_cold = None
        dt_cold = 0
        print(f"    [SBF] cold start ERROR: {e}")

    if res_cold and res_cold.success:
        pl = _path_length(res_cold.path, period)
        gd = res_cold.metadata.get('grow_detail', {})
        result['cold_start'] = {
            'success': True,
            'total_time_ms': dt_cold * 1000,
            'grow_time_ms': res_cold.phase_times.get('grow', 0) * 1000,
            'solve_time_ms': res_cold.phase_times.get('solve', 0) * 1000,
            'path_length': pl,
            'n_boxes': res_cold.nodes_explored,
            'connected': gd.get('connected', False),
            'n_start_boxes': gd.get('n_start_boxes', 0),
            'n_goal_boxes': gd.get('n_goal_boxes', 0),
            'n_random_boxes': gd.get('n_random_boxes', 0),
            'phase_times': dict(res_cold.phase_times),
            'metadata': dict(res_cold.metadata),
        }
    else:
        result['cold_start'] = {
            'success': False,
            'total_time_ms': dt_cold * 1000,
            'error': str(getattr(res_cold, 'metadata', {}).get('error', 'failed')),
        }

    # ── 2) Warm start (with cache) ──
    print("    [SBF] warm start (with cache) ...")
    sbf_warm = SBFAdapter(method="dijkstra")
    cfg_warm = dict(base_cfg)
    sbf_warm.setup(robot, scene, cfg_warm)
    try:
        t0 = time.perf_counter()
        res_warm = sbf_warm.plan(q_start, q_goal, timeout=timeout)
        dt_warm = time.perf_counter() - t0
    except Exception as e:
        res_warm = None
        dt_warm = 0
        print(f"    [SBF] warm start ERROR: {e}")

    if res_warm and res_warm.success:
        pl = _path_length(res_warm.path, period)
        gd = res_warm.metadata.get('grow_detail', {})
        result['warm_start'] = {
            'success': True,
            'total_time_ms': dt_warm * 1000,
            'grow_time_ms': res_warm.phase_times.get('grow', 0) * 1000,
            'solve_time_ms': res_warm.phase_times.get('solve', 0) * 1000,
            'path_length': pl,
            'n_boxes': res_warm.nodes_explored,
            'connected': gd.get('connected', False),
            'n_start_boxes': gd.get('n_start_boxes', 0),
            'n_goal_boxes': gd.get('n_goal_boxes', 0),
            'n_random_boxes': gd.get('n_random_boxes', 0),
            'phase_times': dict(res_warm.phase_times),
            'metadata': dict(res_warm.metadata),
        }
    else:
        result['warm_start'] = {
            'success': False,
            'total_time_ms': dt_warm * 1000,
        }

    # ── 3) Incremental update: add 1 obstacle, re-plan ──
    print("    [SBF] incremental (+1 obstacle) ...")
    if sbf_warm._prep is not None and res_warm and res_warm.success:
        rng_inc = np.random.default_rng(seed + 9999)
        # Generate a new obstacle in workspace
        if robot.n_joints == 2:
            cx = float(rng_inc.uniform(-1.5, 1.5))
            cy = float(rng_inc.uniform(-1.5, 1.5))
            w, h = float(rng_inc.uniform(0.2, 0.5)), float(rng_inc.uniform(0.2, 0.5))
            new_obs = {
                'min_point': [cx - w/2, cy - h/2],
                'max_point': [cx + w/2, cy + h/2],
                'name': 'new_obs_inc',
            }
        else:
            # Panda workspace
            r = float(rng_inc.uniform(0.2, 0.7))
            theta = float(rng_inc.uniform(-math.pi, math.pi))
            cx, cy = r * math.cos(theta), r * math.sin(theta)
            cz = float(rng_inc.uniform(0.15, 0.85))
            hs = float(rng_inc.uniform(0.06, 0.15))
            new_obs = {
                'min_point': [cx - hs, cy - hs, cz - hs],
                'max_point': [cx + hs, cy + hs, cz + hs],
                'name': 'new_obs_inc',
            }

        try:
            t0 = time.perf_counter()
            inc_stats = sbf_warm.update_scene_incremental(
                scene, added_obstacles=[new_obs], regrow_budget=60,
                rng=rng_inc)
            dt_inc_update = (time.perf_counter() - t0) * 1000

            # Re-plan on updated forest
            sbf_warm._prep = None  # force regrow after incremental
            # Actually, incremental update keeps the prep — we just need
            # to re-plan. But plan() uses _prep, so we need to clear solve
            # cache if any. The adapter re-uses prep, so next plan() goes
            # directly to solve.
            # However, update_scene_incremental modifies prep in-place,
            # so _prep is still set — plan() will go to solve.
            # But we want to re-measure the full plan time.
            # Actually, since _prep is set, plan() will skip grow and go
            # straight to solve. This is the expected "incremental" flow.

            # Re-attach _prep (incremental keeps it)
            # sbf_warm._prep should still be valid after incremental update

            t0 = time.perf_counter()
            res_inc = sbf_warm.plan(q_start, q_goal, timeout=timeout)
            dt_inc_plan = (time.perf_counter() - t0) * 1000

            result['incremental'] = {
                'success': res_inc.success if res_inc else False,
                'update_time_ms': dt_inc_update,
                'replan_time_ms': dt_inc_plan,
                'total_time_ms': dt_inc_update + dt_inc_plan,
                'path_length': _path_length(res_inc.path, period) if res_inc and res_inc.success else float('nan'),
                'inc_stats': {k: v for k, v in inc_stats.items()
                              if not k.startswith('_')},
            }
        except Exception as e:
            print(f"    [SBF] incremental ERROR: {e}")
            result['incremental'] = {'success': False, 'error': str(e)}
    else:
        result['incremental'] = {'success': False, 'error': 'no forest to update'}

    # ── 4) Replan: different start/goal on same forest ──
    print("    [SBF] replan (different query, reuse forest) ...")
    # Use a fresh warm adapter but reuse from prior
    sbf_reuse = SBFAdapter(method="dijkstra")
    cfg_reuse = dict(base_cfg)
    sbf_reuse.setup(robot, scene, cfg_reuse)
    try:
        # First plan builds forest
        t0 = time.perf_counter()
        res_first = sbf_reuse.plan(q_start, q_goal, timeout=timeout)
        dt_first = time.perf_counter() - t0

        # Second plan reuses forest (different query)
        t0 = time.perf_counter()
        res_replan = sbf_reuse.plan(q_start_alt, q_goal_alt, timeout=timeout)
        dt_replan = time.perf_counter() - t0

        result['replan'] = {
            'success': res_replan.success if res_replan else False,
            'first_plan_time_ms': dt_first * 1000,
            'replan_time_ms': dt_replan * 1000,
            'replan_solve_ms': res_replan.phase_times.get('solve', 0) * 1000 if res_replan else 0,
            'path_length': _path_length(res_replan.path, period) if res_replan and res_replan.success else float('nan'),
        }
    except Exception as e:
        print(f"    [SBF] replan ERROR: {e}")
        result['replan'] = {'success': False, 'error': str(e)}

    # ── Detailed breakdown (from warm_start) ──
    if res_warm and res_warm.success:
        meta = res_warm.metadata
        gd = meta.get('grow_detail', {})
        result['breakdown'] = {
            'grow_ms': meta.get('grow_ms', 0),
            'coarsen_ms': meta.get('coarsen_ms', 0),
            'adj_ms': meta.get('adj_ms', 0),
            'bridge_ms': meta.get('bridge_ms', 0),
            'plan_ms': meta.get('plan_ms', 0),
            'post_safe_ms': meta.get('post_safe_ms', 0),
            'plan_graph_ms': meta.get('plan_graph_ms', 0),
            'plan_waypoints_ms': meta.get('plan_waypoints_ms', 0),
            'plan_refine_ms': meta.get('plan_refine_ms', 0),
            'plan_shortcut_ms': meta.get('plan_shortcut_ms', 0),
            'cache_start_ms': meta.get('cache_start_ms', 0),
            'growprep_total_ms': meta.get('growprep_total_ms', 0),
            'repair_ms': meta.get('repair_ms', 0),
            # grow sub-phases
            'warmup_ms': gd.get('warmup_ms', 0),
            'sample_ms': gd.get('sample_ms', 0),
            'boundary_ms': gd.get('boundary_ms', 0),
            'is_occupied_ms': gd.get('is_occupied_ms', 0),
            'probe_ms': gd.get('probe_ms', 0),
            'find_free_box_ms': gd.get('find_free_box_ms', 0),
            'add_box_ms': gd.get('add_box_ms', 0),
            'overhead_ms': gd.get('overhead_ms', 0),
        }
    else:
        result['breakdown'] = {}

    return result


# ═══════════════════════════════════════════════════════════════════
# Run OMPL methods
# ═══════════════════════════════════════════════════════════════════

def run_ompl_method(robot, scene, q_start, q_goal, algorithm, seed,
                    timeout=2.0, is_2dof=False):
    """Run a single OMPL method. Returns result dict."""
    period = _compute_period(robot.joint_limits)

    if is_2dof:
        raw = run_ompl_2dof(q_start, q_goal, robot.joint_limits, scene,
                            algorithm, timeout=timeout, seed=seed)
        if raw is None or not raw.get('success', False):
            return {
                'success': False,
                'algorithm': algorithm,
                'planning_time_ms': 0,
            }
        pl = _path_length(raw.get('path'), period)
        return {
            'success': True,
            'algorithm': algorithm,
            'planning_time_ms': raw['planning_time'] * 1000,
            'first_solution_time_ms': raw.get('first_solution_time', 0) * 1000,
            'first_solution_cost': _safe_float(raw.get('first_solution_cost')),
            'path_length': pl,
            'cost_history': raw.get('cost_history', []),
            'n_nodes': raw.get('n_nodes', 0),
        }
    else:
        # Panda: use OMPLPlanner adapter
        planner = OMPLPlanner(algorithm=algorithm)
        planner.setup(robot, scene, {'seed': seed, 'step_size': 0.5})
        res = planner.plan(q_start, q_goal, timeout=timeout)
        if not res.success:
            return {
                'success': False,
                'algorithm': algorithm,
                'planning_time_ms': res.planning_time * 1000,
            }
        pl = _path_length(res.path, period)
        meta = res.metadata or {}
        return {
            'success': True,
            'algorithm': algorithm,
            'planning_time_ms': res.planning_time * 1000,
            'first_solution_time_ms': _safe_float(res.first_solution_time) * 1000,
            'first_solution_cost': _safe_float(meta.get('first_solution_cost')),
            'path_length': pl,
            'cost_history': meta.get('cost_history', []),
            'n_nodes': res.nodes_explored,
        }


# ═══════════════════════════════════════════════════════════════════
# Per-seed experiment
# ═══════════════════════════════════════════════════════════════════

def run_one_seed(robot, scene, q_start, q_goal,
                 q_start_alt, q_goal_alt,
                 seed, timeout_ompl=2.0, timeout_sbf=30.0,
                 is_2dof=False):
    """Run all methods on one scene. Returns dict."""
    import copy
    period = _compute_period(robot.joint_limits)
    sbf_cfg = SBF_CFG_2DOF if is_2dof else SBF_CFG_PANDA

    # 保存原始场景用于 OMPL (SBF incremental 会修改 scene — 添加障碍物)
    scene_for_ompl = copy.deepcopy(scene)

    print(f"\n  seed={seed}")

    # ── SBF-Dijkstra (detailed) ──
    sbf = run_sbf_detailed(robot, scene, q_start, q_goal, seed,
                           q_start_alt, q_goal_alt, timeout=timeout_sbf,
                           sbf_cfg_override=sbf_cfg)

    # ── OMPL RRT-Connect (使用原始场景，不受 incremental 影响) ──
    print("    [OMPL] RRTConnect ...")
    rrt = run_ompl_method(robot, scene_for_ompl, q_start, q_goal,
                          "RRTConnect", seed,
                          timeout=timeout_ompl, is_2dof=is_2dof)

    # ── OMPL BIT* ──
    print("    [OMPL] BITstar ...")
    bit = run_ompl_method(robot, scene_for_ompl, q_start, q_goal,
                          "BITstar", seed,
                          timeout=timeout_ompl, is_2dof=is_2dof)

    # ── BIT* time to reach SBF quality ──
    sbf_cost = float('nan')
    if sbf.get('warm_start', {}).get('success'):
        sbf_cost = sbf['warm_start'].get('path_length', float('nan'))

    bit_to_sbf = None
    if bit.get('success') and not math.isnan(sbf_cost):
        bit_to_sbf = _find_time_to_reach_cost(
            bit.get('cost_history', []),
            sbf_cost,
            final_cost=bit.get('path_length'),
            total_time=bit.get('planning_time_ms', 0) / 1000)
        if bit_to_sbf is not None:
            bit_to_sbf *= 1000  # convert to ms

    # Summary print
    ws = sbf.get('warm_start', {})
    print(f"    SBF:  {'OK' if ws.get('success') else 'FAIL'}  "
          f"{ws.get('total_time_ms', 0):.0f}ms  "
          f"path={ws.get('path_length', 'N/A')}")
    print(f"    RRT:  {'OK' if rrt.get('success') else 'FAIL'}  "
          f"{rrt.get('planning_time_ms', 0):.0f}ms  "
          f"path={rrt.get('path_length', 'N/A')}")
    print(f"    BIT*: {'OK' if bit.get('success') else 'FAIL'}  "
          f"{bit.get('planning_time_ms', 0):.0f}ms  "
          f"path={bit.get('path_length', 'N/A')}  "
          f"to_sbf={'%.0fms' % bit_to_sbf if bit_to_sbf else 'N/A'}")

    return {
        'seed': seed,
        'sbf': sbf,
        'rrt': rrt,
        'bit': bit,
        'bit_to_sbf_ms': bit_to_sbf,
    }


# ═══════════════════════════════════════════════════════════════════
# Excel output
# ═══════════════════════════════════════════════════════════════════

def write_excel(results_2d: List[dict], results_panda: List[dict],
                out_path: Path):
    """Write comparison results to Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ── Style helpers ──
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill(start_color="4472C4", end_color="4472C4",
                           fill_type="solid")
    hdr_align = Alignment(horizontal="center", vertical="center",
                          wrap_text=True)
    even_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3",
                            fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="B4C6E7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    avg_font = Font(bold=True, size=11)
    avg_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC",
                           fill_type="solid")

    def _style_header(ws, row, ncols):
        for col in range(1, ncols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = hdr_align

    def _style_data(ws, start_row, end_row, ncols):
        for r in range(start_row, end_row + 1):
            for c in range(1, ncols + 1):
                cell = ws.cell(row=r, column=c)
                cell.alignment = center_align
                cell.border = border
                if (r - start_row) % 2 == 1:
                    cell.fill = even_fill

    def _style_avg_row(ws, row, ncols):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = avg_font
            cell.fill = avg_fill

    def _avg(lst):
        clean = [x for x in lst if x is not None and not math.isnan(x)]
        return sum(clean) / len(clean) if clean else float('nan')

    def _median(lst):
        clean = sorted(x for x in lst if x is not None and not math.isnan(x))
        if not clean:
            return float('nan')
        n = len(clean)
        return (clean[n//2] + clean[(n-1)//2]) / 2

    def _fmt(v, decimals=1):
        if v is None or (isinstance(v, float) and math.isnan(v)):
            return "N/A"
        return round(v, decimals)

    # ═══════════════════════════════════════════════════════════
    # Sheet: summary
    # ═══════════════════════════════════════════════════════════
    ws_sum = wb.active
    ws_sum.title = "summary"

    sum_headers = [
        "Environment", "Method",
        "Success\nRate", "Mean Time\n(ms)", "Median Time\n(ms)",
        "P90 Time\n(ms)", "Mean Path\nLength",
        "BIT* Time\nto SBF (ms)",
    ]
    ncols = len(sum_headers)
    for col, h in enumerate(sum_headers, 1):
        ws_sum.cell(row=1, column=col, value=h)
    _style_header(ws_sum, 1, ncols)

    row_sum = 2
    for env_name, env_results in [("2dof", results_2d), ("panda", results_panda)]:
        if not env_results:
            continue
        # SBF summary (warm start)
        sbf_times = [r['sbf'].get('warm_start', {}).get('total_time_ms', float('nan'))
                     for r in env_results if r['sbf'].get('warm_start', {}).get('success')]
        sbf_paths = [r['sbf'].get('warm_start', {}).get('path_length', float('nan'))
                     for r in env_results if r['sbf'].get('warm_start', {}).get('success')]
        n_sbf_ok = len(sbf_times)
        n_total = len(env_results)

        ws_sum.cell(row=row_sum, column=1, value=env_name)
        ws_sum.cell(row=row_sum, column=2, value="SBF-Dijkstra")
        ws_sum.cell(row=row_sum, column=3, value=f"{n_sbf_ok}/{n_total}")
        ws_sum.cell(row=row_sum, column=4, value=_fmt(_avg(sbf_times)))
        ws_sum.cell(row=row_sum, column=5, value=_fmt(_median(sbf_times)))
        ws_sum.cell(row=row_sum, column=6,
                    value=_fmt(float(np.percentile(sbf_times, 90)) if sbf_times else float('nan')))
        ws_sum.cell(row=row_sum, column=7, value=_fmt(_avg(sbf_paths), 3))
        ws_sum.cell(row=row_sum, column=8, value="—")
        row_sum += 1

        # RRT summary
        rrt_times = [r['rrt'].get('planning_time_ms', float('nan'))
                     for r in env_results if r['rrt'].get('success')]
        rrt_paths = [r['rrt'].get('path_length', float('nan'))
                     for r in env_results if r['rrt'].get('success')]
        n_rrt_ok = len(rrt_times)

        ws_sum.cell(row=row_sum, column=1, value=env_name)
        ws_sum.cell(row=row_sum, column=2, value="OMPL-RRTConnect")
        ws_sum.cell(row=row_sum, column=3, value=f"{n_rrt_ok}/{n_total}")
        ws_sum.cell(row=row_sum, column=4, value=_fmt(_avg(rrt_times)))
        ws_sum.cell(row=row_sum, column=5, value=_fmt(_median(rrt_times)))
        ws_sum.cell(row=row_sum, column=6,
                    value=_fmt(float(np.percentile(rrt_times, 90)) if rrt_times else float('nan')))
        ws_sum.cell(row=row_sum, column=7, value=_fmt(_avg(rrt_paths), 3))
        ws_sum.cell(row=row_sum, column=8, value="—")
        row_sum += 1

        # BIT* summary
        bit_times = [r['bit'].get('planning_time_ms', float('nan'))
                     for r in env_results if r['bit'].get('success')]
        bit_paths = [r['bit'].get('path_length', float('nan'))
                     for r in env_results if r['bit'].get('success')]
        bit_to_sbf = [r['bit_to_sbf_ms'] for r in env_results
                      if r.get('bit_to_sbf_ms') is not None]
        n_bit_ok = len(bit_times)

        ws_sum.cell(row=row_sum, column=1, value=env_name)
        ws_sum.cell(row=row_sum, column=2, value="OMPL-BIT*")
        ws_sum.cell(row=row_sum, column=3, value=f"{n_bit_ok}/{n_total}")
        ws_sum.cell(row=row_sum, column=4, value=_fmt(_avg(bit_times)))
        ws_sum.cell(row=row_sum, column=5, value=_fmt(_median(bit_times)))
        ws_sum.cell(row=row_sum, column=6,
                    value=_fmt(float(np.percentile(bit_times, 90)) if bit_times else float('nan')))
        ws_sum.cell(row=row_sum, column=7, value=_fmt(_avg(bit_paths), 3))
        ws_sum.cell(row=row_sum, column=8,
                    value=f"{_fmt(_avg(bit_to_sbf))} ({len(bit_to_sbf)}/{n_total})"
                    if bit_to_sbf else "N/A")
        row_sum += 1

    _style_data(ws_sum, 2, row_sum - 1, ncols)
    for i in range(1, ncols + 1):
        ws_sum.column_dimensions[get_column_letter(i)].width = 18
    ws_sum.freeze_panes = "C2"

    # ═══════════════════════════════════════════════════════════
    # Per-env sheets: 2dof / panda
    # ═══════════════════════════════════════════════════════════
    def _write_env_sheet(ws, env_results, env_name):
        headers = [
            "Seed",
            "SBF Cold\n(ms)", "SBF Warm\n(ms)", "SBF Warm\nPath",
            "SBF Inc\n(ms)", "SBF Replan\n(ms)",
            "SBF #Boxes", "SBF Start\nBoxes", "SBF Goal\nBoxes",
            "SBF\nConnected",
            "RRT Time\n(ms)", "RRT Path",
            "BIT* Time\n(ms)", "BIT* Path",
            "BIT* 1st Sol\n(ms)", "BIT* to SBF\n(ms)",
        ]
        ncols = len(headers)
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h)
        _style_header(ws, 1, ncols)

        row = 2
        accum = {k: [] for k in [
            'cold', 'warm', 'warm_path', 'inc', 'replan',
            'n_boxes', 'start_boxes', 'goal_boxes',
            'rrt_time', 'rrt_path', 'bit_time', 'bit_path',
            'bit_1st', 'bit_to_sbf']}

        for r in env_results:
            ws.cell(row=row, column=1, value=r['seed'])

            sbf = r['sbf']
            cold = sbf.get('cold_start', {})
            warm = sbf.get('warm_start', {})
            inc = sbf.get('incremental', {})
            replan = sbf.get('replan', {})

            cold_ms = cold.get('total_time_ms', float('nan'))
            warm_ms = warm.get('total_time_ms', float('nan'))
            warm_path = warm.get('path_length', float('nan'))
            inc_ms = inc.get('total_time_ms', float('nan')) if inc.get('success') else float('nan')
            replan_ms = replan.get('replan_time_ms', float('nan')) if replan.get('success') else float('nan')
            n_boxes = warm.get('n_boxes', 0) if warm.get('success') else 0
            start_boxes = warm.get('n_start_boxes', 0)
            goal_boxes = warm.get('n_goal_boxes', 0)
            connected = warm.get('connected', False)

            ws.cell(row=row, column=2, value=_fmt(cold_ms) if cold.get('success') else "FAIL")
            ws.cell(row=row, column=3, value=_fmt(warm_ms) if warm.get('success') else "FAIL")
            ws.cell(row=row, column=4, value=_fmt(warm_path, 3) if warm.get('success') else "FAIL")
            ws.cell(row=row, column=5, value=_fmt(inc_ms) if inc.get('success') else "FAIL")
            ws.cell(row=row, column=6, value=_fmt(replan_ms) if replan.get('success') else "FAIL")
            ws.cell(row=row, column=7, value=n_boxes)
            ws.cell(row=row, column=8, value=start_boxes)
            ws.cell(row=row, column=9, value=goal_boxes)
            ws.cell(row=row, column=10, value="Y" if connected else "N")

            rrt = r['rrt']
            ws.cell(row=row, column=11,
                    value=_fmt(rrt['planning_time_ms']) if rrt.get('success') else "FAIL")
            ws.cell(row=row, column=12,
                    value=_fmt(rrt.get('path_length'), 3) if rrt.get('success') else "FAIL")

            bit = r['bit']
            ws.cell(row=row, column=13,
                    value=_fmt(bit['planning_time_ms']) if bit.get('success') else "FAIL")
            ws.cell(row=row, column=14,
                    value=_fmt(bit.get('path_length'), 3) if bit.get('success') else "FAIL")
            ws.cell(row=row, column=15,
                    value=_fmt(_safe_float(bit.get('first_solution_time_ms'))))
            ws.cell(row=row, column=16,
                    value=_fmt(r.get('bit_to_sbf_ms')) if r.get('bit_to_sbf_ms') else "N/A")

            # accumulate
            if cold.get('success'):
                accum['cold'].append(cold_ms)
            if warm.get('success'):
                accum['warm'].append(warm_ms)
                accum['warm_path'].append(warm_path)
                accum['n_boxes'].append(n_boxes)
                accum['start_boxes'].append(start_boxes)
                accum['goal_boxes'].append(goal_boxes)
            if inc.get('success'):
                accum['inc'].append(inc_ms)
            if replan.get('success'):
                accum['replan'].append(replan_ms)
            if rrt.get('success'):
                accum['rrt_time'].append(rrt['planning_time_ms'])
                accum['rrt_path'].append(rrt.get('path_length', float('nan')))
            if bit.get('success'):
                accum['bit_time'].append(bit['planning_time_ms'])
                accum['bit_path'].append(bit.get('path_length', float('nan')))
                accum['bit_1st'].append(_safe_float(bit.get('first_solution_time_ms')))
            if r.get('bit_to_sbf_ms') is not None:
                accum['bit_to_sbf'].append(r['bit_to_sbf_ms'])

            row += 1

        _style_data(ws, 2, row - 1, ncols)

        # Average row
        ws.cell(row=row, column=1, value="AVG")
        ws.cell(row=row, column=2, value=_fmt(_avg(accum['cold'])))
        ws.cell(row=row, column=3, value=_fmt(_avg(accum['warm'])))
        ws.cell(row=row, column=4, value=_fmt(_avg(accum['warm_path']), 3))
        ws.cell(row=row, column=5, value=_fmt(_avg(accum['inc'])))
        ws.cell(row=row, column=6, value=_fmt(_avg(accum['replan'])))
        ws.cell(row=row, column=7, value=_fmt(_avg(accum['n_boxes']), 0))
        ws.cell(row=row, column=8, value=_fmt(_avg(accum['start_boxes']), 0))
        ws.cell(row=row, column=9, value=_fmt(_avg(accum['goal_boxes']), 0))
        ws.cell(row=row, column=10, value="")
        ws.cell(row=row, column=11, value=_fmt(_avg(accum['rrt_time'])))
        ws.cell(row=row, column=12, value=_fmt(_avg(accum['rrt_path']), 3))
        ws.cell(row=row, column=13, value=_fmt(_avg(accum['bit_time'])))
        ws.cell(row=row, column=14, value=_fmt(_avg(accum['bit_path']), 3))
        ws.cell(row=row, column=15, value=_fmt(_avg(accum['bit_1st'])))
        n_reached = len(accum['bit_to_sbf'])
        ws.cell(row=row, column=16,
                value=f"{_fmt(_avg(accum['bit_to_sbf']))} ({n_reached}/{len(env_results)})"
                if accum['bit_to_sbf'] else "N/A")

        _style_avg_row(ws, row, ncols)

        for i in range(1, ncols + 1):
            ws.column_dimensions[get_column_letter(i)].width = 16
        ws.freeze_panes = "B2"

    # ═══════════════════════════════════════════════════════════
    # Sheet: sbf_breakdown
    # ═══════════════════════════════════════════════════════════
    def _write_breakdown_sheet(ws, results_2d, results_panda):
        bd_keys = [
            ('grow_ms', 'Grow (ms)'),
            ('coarsen_ms', 'Coarsen (ms)'),
            ('adj_ms', 'Adj (ms)'),
            ('bridge_ms', 'Bridge (ms)'),
            ('plan_ms', 'Plan (ms)'),
            ('plan_graph_ms', 'Plan Graph (ms)'),
            ('plan_waypoints_ms', 'Plan WP (ms)'),
            ('plan_refine_ms', 'Plan Refine (ms)'),
            ('plan_shortcut_ms', 'Plan Shortcut (ms)'),
            ('post_safe_ms', 'Post Safe (ms)'),
            ('warmup_ms', 'Warmup (ms)'),
            ('sample_ms', 'Sample (ms)'),
            ('boundary_ms', 'Boundary (ms)'),
            ('is_occupied_ms', 'IsOccupied (ms)'),
            ('probe_ms', 'Probe (ms)'),
            ('find_free_box_ms', 'FindFreeBox (ms)'),
            ('add_box_ms', 'AddBox (ms)'),
            ('overhead_ms', 'Overhead (ms)'),
        ]
        headers = ["Env", "Seed"] + [h for _, h in bd_keys]
        ncols = len(headers)
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h)
        _style_header(ws, 1, ncols)

        row = 2
        for env_name, env_results in [("2dof", results_2d), ("panda", results_panda)]:
            accum = {k: [] for k, _ in bd_keys}
            for r in env_results:
                bd = r['sbf'].get('breakdown', {})
                if not bd:
                    continue
                ws.cell(row=row, column=1, value=env_name)
                ws.cell(row=row, column=2, value=r['seed'])
                for ci, (k, _) in enumerate(bd_keys):
                    v = bd.get(k, 0)
                    ws.cell(row=row, column=3 + ci, value=round(v, 2))
                    accum[k].append(v)
                row += 1

            # Avg row for this env
            if accum[bd_keys[0][0]]:
                ws.cell(row=row, column=1, value=f"{env_name} AVG")
                ws.cell(row=row, column=2, value="")
                for ci, (k, _) in enumerate(bd_keys):
                    ws.cell(row=row, column=3 + ci,
                            value=round(_avg(accum[k]), 2))
                _style_avg_row(ws, row, ncols)
                row += 1

        _style_data(ws, 2, row - 1, ncols)
        for i in range(1, ncols + 1):
            ws.column_dimensions[get_column_letter(i)].width = 16
        ws.freeze_panes = "C2"

    # ═══════════════════════════════════════════════════════════
    # Sheet: abnormal_cases
    # ═══════════════════════════════════════════════════════════
    def _write_abnormal_sheet(ws, results_2d, results_panda):
        headers = ["Env", "Seed", "Method", "Issue", "Details"]
        ncols = len(headers)
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h)
        _style_header(ws, 1, ncols)

        row = 2
        for env_name, env_results in [("2dof", results_2d), ("panda", results_panda)]:
            for r in env_results:
                seed = r['seed']
                # SBF failures
                for mode in ['cold_start', 'warm_start', 'incremental', 'replan']:
                    sub = r['sbf'].get(mode, {})
                    if not sub.get('success', True):
                        ws.cell(row=row, column=1, value=env_name)
                        ws.cell(row=row, column=2, value=seed)
                        ws.cell(row=row, column=3, value=f"SBF-{mode}")
                        ws.cell(row=row, column=4, value="FAIL")
                        ws.cell(row=row, column=5,
                                value=sub.get('error', ''))
                        row += 1
                # OMPL failures
                for method_key, method_name in [('rrt', 'RRTConnect'), ('bit', 'BIT*')]:
                    m = r.get(method_key, {})
                    if not m.get('success', True):
                        ws.cell(row=row, column=1, value=env_name)
                        ws.cell(row=row, column=2, value=seed)
                        ws.cell(row=row, column=3, value=method_name)
                        ws.cell(row=row, column=4, value="FAIL")
                        ws.cell(row=row, column=5, value="")
                        row += 1

        _style_data(ws, 2, max(row - 1, 2), ncols)
        for i in range(1, ncols + 1):
            ws.column_dimensions[get_column_letter(i)].width = 20
        ws.freeze_panes = "A2"

    # ── Build sheets ──
    _write_env_sheet(wb.create_sheet("2dof"), results_2d, "2dof")
    _write_env_sheet(wb.create_sheet("panda"), results_panda, "panda")
    _write_breakdown_sheet(wb.create_sheet("sbf_breakdown"),
                           results_2d, results_panda)
    _write_abnormal_sheet(wb.create_sheet("abnormal_cases"),
                          results_2d, results_panda)

    wb.save(out_path)
    print(f"\nExcel saved: {out_path}")


# ═══════════════════════════════════════════════════════════════════
# Main
# ═══════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        description="Exp2: SBF-Dijkstra vs OMPL method comparison")
    parser.add_argument("--seeds", type=int, default=30,
                        help="Number of random seeds (default: 30)")
    parser.add_argument("--start-seed", type=int, default=0,
                        help="Starting seed (default: 0)")
    parser.add_argument("--timeout-ompl", type=float, default=10.0,
                        help="OMPL timeout per query (s)")
    parser.add_argument("--timeout-sbf", type=float, default=30.0,
                        help="SBF planning timeout (s)")
    parser.add_argument("--scene-seed", type=int, default=1000,
                        help="Fixed seed for scene generation (default: 1000, same as exp1)")
    parser.add_argument("--skip-2dof", action="store_true")
    parser.add_argument("--skip-panda", action="store_true")
    args = parser.parse_args()

    seeds = list(range(args.start_seed, args.start_seed + args.seeds))
    print(f"Exp2: Method comparison")
    print(f"  Seeds: {args.seeds} ({args.start_seed}..{args.start_seed + args.seeds - 1})")
    print(f"  OMPL timeout: {args.timeout_ompl}s")

    # ── 2DOF ──
    results_2d = []
    if not args.skip_2dof:
        print(f"\n{'='*60}")
        print(f"  2DOF (5 obstacles)")
        print(f"{'='*60}")

        robot_2d = load_robot("2dof_planar")
        q_start_2d = np.array([-2.0, 1.5])
        q_goal_2d  = np.array([2.0, -1.5])
        # Alternative query for replan test
        q_start_alt_2d = np.array([1.0, -1.0])
        q_goal_alt_2d  = np.array([-1.0, 1.0])

        for s in seeds:
            rng = np.random.default_rng(args.scene_seed + s)
            try:
                scene = build_random_2d_scene(robot_2d, q_start_2d, q_goal_2d,
                                              rng, n_obstacles=5)
            except RuntimeError:
                print(f"  seed={s}: SKIP (no valid scene)")
                continue

            r = run_one_seed(robot_2d, scene, q_start_2d, q_goal_2d,
                             q_start_alt_2d, q_goal_alt_2d,
                             seed=s, timeout_ompl=args.timeout_ompl,
                             timeout_sbf=args.timeout_sbf,
                             is_2dof=True)
            results_2d.append(r)

    # ── Panda ──
    results_panda = []
    if not args.skip_panda:
        print(f"\n{'='*60}")
        print(f"  Panda 7-DOF (5 obstacles)")
        print(f"{'='*60}")

        robot_panda = load_robot("panda")
        q_start_panda = np.array([0.5, -1.2, 0.5, -2.5, 0.5, 0.8, 1.5])
        q_goal_panda  = np.array([-2.0, 1.2, -1.8, -0.5, -2.0, 3.0, -1.8])
        # Alternative query
        q_start_alt_panda = np.array([0.0, 0.0, 0.0, -1.57, 0.0, 1.57, 0.785])
        q_goal_alt_panda  = np.array([-1.5, 0.8, -1.0, -0.8, -1.5, 2.5, -1.0])

        for s in seeds:
            try:
                scene = build_panda_5obs_scene(
                    robot_panda, q_start_panda, q_goal_panda,
                    scene_seed=args.scene_seed + s)
            except RuntimeError:
                print(f"  seed={s}: SKIP (no valid scene)")
                continue

            r = run_one_seed(robot_panda, scene,
                             q_start_panda, q_goal_panda,
                             q_start_alt_panda, q_goal_alt_panda,
                             seed=s, timeout_ompl=args.timeout_ompl,
                             timeout_sbf=args.timeout_sbf,
                             is_2dof=False)
            results_panda.append(r)

    # ── Save JSON ──
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_dir = _ROOT / "experiments" / "output"
    out_dir.mkdir(parents=True, exist_ok=True)

    # Serialize — strip numpy arrays for JSON
    def _clean(obj):
        if isinstance(obj, np.ndarray):
            return obj.tolist()
        if isinstance(obj, dict):
            return {k: _clean(v) for k, v in obj.items()}
        if isinstance(obj, (list, tuple)):
            return [_clean(v) for v in obj]
        if isinstance(obj, (np.integer,)):
            return int(obj)
        if isinstance(obj, (np.floating,)):
            return float(obj)
        return obj

    json_path = out_dir / f"exp2_comparison_{ts}.json"
    raw_data = {
        'timestamp': ts,
        'seeds': seeds,
        'timeout_ompl': args.timeout_ompl,
        '2dof': _clean(results_2d),
        'panda': _clean(results_panda),
    }
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(raw_data, f, indent=2, default=str)
    print(f"\nJSON saved: {json_path}")

    # ── Excel ──
    xlsx_path = out_dir / f"exp2_method_comparison_{ts}.xlsx"
    write_excel(results_2d, results_panda, xlsx_path)

    # ── Console summary ──
    print(f"\n{'='*60}")
    print("  SUMMARY")
    print(f"{'='*60}")
    for env_name, env_results in [("2dof", results_2d), ("panda", results_panda)]:
        if not env_results:
            continue
        n = len(env_results)
        sbf_ok = sum(1 for r in env_results
                     if r['sbf'].get('warm_start', {}).get('success'))
        rrt_ok = sum(1 for r in env_results if r['rrt'].get('success'))
        bit_ok = sum(1 for r in env_results if r['bit'].get('success'))
        print(f"\n  {env_name} ({n} seeds):")
        print(f"    SBF success:  {sbf_ok}/{n}")
        print(f"    RRT success:  {rrt_ok}/{n}")
        print(f"    BIT* success: {bit_ok}/{n}")

        sbf_times = [r['sbf']['warm_start']['total_time_ms']
                     for r in env_results
                     if r['sbf'].get('warm_start', {}).get('success')]
        if sbf_times:
            print(f"    SBF median:   {np.median(sbf_times):.0f} ms")
        rrt_times = [r['rrt']['planning_time_ms']
                     for r in env_results if r['rrt'].get('success')]
        if rrt_times:
            print(f"    RRT median:   {np.median(rrt_times):.0f} ms")


if __name__ == "__main__":
    main()
