"""Generate benchmark_comparison_v5.xlsx from raw JSON results.

Reads: experiments/output/raw/panda_10obs_comparison.json
Writes: experiments/output/benchmark_comparison_v5.xlsx
"""
from __future__ import annotations
import json, sys
from pathlib import Path

import numpy as np

_ROOT = Path(__file__).resolve().parents[1]
_SRC = _ROOT / "src"
for p in (_ROOT, _SRC):
    if str(p) not in sys.path:
        sys.path.insert(0, str(p))

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("pip install openpyxl")
    sys.exit(1)


def _find_time_to_quality(cost_history: list, target_raw_cost: float):
    """Find earliest time in cost_history where cost <= target."""
    for t, c in cost_history:
        if c <= target_raw_cost:
            return t
    return None


def main():
    raw_path = _ROOT / "experiments" / "output" / "raw" / "panda_10obs_comparison.json"
    d = json.load(open(raw_path, encoding="utf-8"))

    reuse = d.get("reuse_results", [])

    # ── Collect per-seed data ──
    seeds = sorted(set(r["seed"] for r in reuse))

    # SBF first query
    sbf_first = {}
    for r in reuse:
        if r["query_index"] == 0 and r.get("success"):
            sbf_first[r["seed"]] = r

    # SBF reuse (query_index > 0)
    sbf_reuse = {}
    for r in reuse:
        if r["query_index"] > 0 and r.get("success"):
            sbf_reuse.setdefault(r["seed"], []).append(r)

    # OMPL results
    rrtc = {}
    bitstar = {}
    for t in d["results"]:
        if t["planner"] == "OMPL-RRTConnect" and t.get("success"):
            rrtc[t["seed"]] = t
        elif t["planner"] == "OMPL-BITstar" and t.get("success"):
            bitstar[t["seed"]] = t

    # ── Create workbook ──
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Panda (10-obs)"

    # Styles
    header_font = Font(bold=True, size=10)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    num_align = Alignment(horizontal="center", vertical="center")
    avg_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"))

    # Headers (matching v4 format)
    headers = [
        ("Seed", 6),
        ("SBF Forest\nBuild (s)", 12),
        ("SBF Query\n(solve) (s)", 12),
        ("SBF Total\nTime (s)", 10),
        ("SBF\nPath Length", 12),
        ("SBF Reuse\nQuery Avg (s)", 13),
        ("RRT-Connect\n1st Sol (s)", 12),
        ("RRT-Connect\nPath Length", 13),
        ("BIT*\nTotal Time (s)", 12),
        ("BIT*\n1st Sol Time (s)", 13),
        ("BIT*\n1st Sol Cost", 12),
        ("BIT*\nFinal Path Length", 14),
        ("BIT*\nTime to SBF\nQuality (s)", 14),
        ("BIT* Reached\nSBF Quality?", 12),
        ("Note", 20),
    ]

    for ci, (hdr, width) in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=ci, value=hdr)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(ci)].width = width

    ws.row_dimensions[1].height = 45

    # ── Data rows ──
    row = 2
    all_grow, all_solve, all_total, all_sbf_cost = [], [], [], []
    all_sbf_reuse = []
    all_rrtc_time, all_rrtc_cost = [], []
    all_bitstar_time, all_bitstar_first, all_bitstar_cost, all_bitstar_final = [], [], [], []
    n_reached_sbf = 0
    time_to_sbf_values = []

    for seed in seeds:
        notes = []

        # SBF
        sf = sbf_first.get(seed)
        if sf:
            grow_s = sf["phase_times"]["grow"]
            solve_s = sf["phase_times"]["solve"]
            total_s = sf["wall_clock"]
            sbf_cost = sf["cost"]
            all_grow.append(grow_s)
            all_solve.append(solve_s)
            all_total.append(total_s)
            all_sbf_cost.append(sbf_cost)
        else:
            grow_s = solve_s = total_s = sbf_cost = None

        # SBF reuse avg
        sr = sbf_reuse.get(seed, [])
        if sr:
            reuse_avg = np.mean([r["phase_times"]["solve"] for r in sr])
            all_sbf_reuse.append(reuse_avg)
        else:
            reuse_avg = None

        # RRT-Connect: use first_solution_time (actual solve, not subprocess overhead)
        rc = rrtc.get(seed)
        if rc:
            rrtc_time = rc.get("first_solution_time", rc["planning_time"])
            rrtc_cost = rc["cost"]
            all_rrtc_time.append(rrtc_time)
            all_rrtc_cost.append(rrtc_cost)
            # Flag outliers
            if rc["planning_time"] > 10:
                notes.append("RRT subprocess outlier (%.1fs)" % rc["planning_time"])
        else:
            rrtc_time = rrtc_cost = None

        # BIT*
        bs = bitstar.get(seed)
        if bs:
            bs_time = bs["planning_time"]
            bs_first_time = bs.get("first_solution_time", bs_time)
            bs_first_cost = bs.get("first_solution_cost", None)
            bs_final_cost = bs["cost"]
            all_bitstar_time.append(bs_time)
            all_bitstar_first.append(bs_first_time)
            all_bitstar_cost.append(bs_final_cost)

            # Time to reach SBF quality
            cost_history = bs.get("cost_history", [])
            time_to_sbf = None
            reached_sbf = "No"
            if sbf_cost is not None and cost_history:
                # cost_history is in OMPL raw space; convert SBF target
                raw_final = bs.get("raw_path_length", bs_final_cost)
                if raw_final > 0 and bs_final_cost > 0:
                    ratio = raw_final / bs_final_cost
                    target_raw = sbf_cost * ratio
                    time_to_sbf = _find_time_to_quality(cost_history, target_raw)
                if time_to_sbf is not None:
                    reached_sbf = "Yes"
                    n_reached_sbf += 1
                    time_to_sbf_values.append(time_to_sbf)
                    # Check if final cost was already better
                    if bs_final_cost < sbf_cost:
                        notes.append("geodesic < total time")
        else:
            bs_time = bs_first_time = bs_first_cost = bs_final_cost = None
            time_to_sbf = None
            reached_sbf = ""

        # Write row
        vals = [
            seed,
            round(grow_s, 3) if grow_s is not None else "",
            round(solve_s, 3) if solve_s is not None else "",
            round(total_s, 3) if total_s is not None else "",
            round(sbf_cost, 3) if sbf_cost is not None else "",
            round(reuse_avg, 3) if reuse_avg is not None else "",
            round(rrtc_time, 3) if rrtc_time is not None else "",
            round(rrtc_cost, 3) if rrtc_cost is not None else "",
            round(bs_time, 2) if bs_time is not None else "",
            round(bs_first_time, 3) if bs_first_time is not None else "",
            round(bs_first_cost, 3) if bs_first_cost is not None else "",
            round(bs_final_cost, 3) if bs_final_cost is not None else "",
            round(time_to_sbf, 2) if time_to_sbf is not None else ("N/A (>timeout)" if reached_sbf == "No" else ""),
            reached_sbf,
            "; ".join(notes) if notes else "",
        ]
        for ci, v in enumerate(vals, start=1):
            cell = ws.cell(row=row, column=ci, value=v)
            cell.alignment = num_align
            cell.border = thin_border
        row += 1

    # ── AVG row ──
    avg_row = row
    def _avg(lst):
        return round(float(np.mean(lst)), 3) if lst else ""

    n_seeds = len(seeds)
    avg_vals = [
        "AVG",
        _avg(all_grow),
        _avg(all_solve),
        _avg(all_total),
        _avg(all_sbf_cost),
        _avg(all_sbf_reuse),
        _avg(all_rrtc_time),
        _avg(all_rrtc_cost),
        round(np.mean(all_bitstar_time), 2) if all_bitstar_time else "",
        _avg(all_bitstar_first),
        "",  # 1st sol cost avg not very meaningful
        _avg(all_bitstar_cost),
        "%.2f (%d/%d)" % (np.mean(time_to_sbf_values), n_reached_sbf, n_seeds) if time_to_sbf_values else "N/A",
        "%d/%d" % (n_reached_sbf, n_seeds),
        "",
    ]
    for ci, v in enumerate(avg_vals, start=1):
        cell = ws.cell(row=avg_row, column=ci, value=v)
        cell.alignment = num_align
        cell.font = Font(bold=True, size=10)
        cell.fill = avg_fill
        cell.border = thin_border

    # ── Config info sheet ──
    ws2 = wb.create_sheet("Config")
    config_info = [
        ("Parameter", "Value"),
        ("Scene", "panda_10obs_moderate"),
        ("Obstacle seed", 200),
        ("n_obstacles", 10),
        ("Robot", "Panda 7-DOF"),
        ("ffb_min_edge", 0.04),
        ("max_boxes", 2000),
        ("RRT bridge timeout (ms)", 200),
        ("OMPL timeout (s)", 30),
        ("Seeds", "0-4"),
        ("Date", d.get("metadata", {}).get("timestamp", "N/A")),
        ("", ""),
        ("Note: RRT-Connect 'Time' column", "Reports first_solution_time (actual OMPL"),
        ("", "solve time), NOT subprocess wall clock."),
        ("", "RRTConnect is single-shot: first sol = only sol."),
        ("", "Subprocess overhead (WSL+Python+OMPL init) adds ~1-10s."),
    ]
    for ri, (k, v) in enumerate(config_info, start=1):
        ws2.cell(row=ri, column=1, value=k).font = Font(bold=(ri == 1))
        ws2.cell(row=ri, column=2, value=v)
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 50

    # Save
    out = _ROOT / "experiments" / "output" / "benchmark_comparison_v5.xlsx"
    wb.save(out)
    print("Saved:", out)


if __name__ == "__main__":
    main()
