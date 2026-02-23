"""
generate_excel_report.py — 生成实验对比 Excel 报告

Sheet 1: 2DOF (8-obs) — SBF-Dijkstra vs RRTConnect vs BIT*
  每个 seed: SBF time/cost, RRTConnect time/cost, BIT* time/cost,
             BIT* 达到 SBF 路径质量所需时间
  底部: 平均值行

Sheet 2: Panda (10-obs) — SBF-Dijkstra vs RRTConnect vs BIT*
  每个 seed: 同上 + 路径质量比较
"""
from __future__ import annotations

import json
import math
from pathlib import Path

import numpy as np

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
from openpyxl.utils import get_column_letter


def _find_time_to_reach_cost(cost_history: list, target_cost: float,
                             final_cost: float | None = None,
                             total_time: float | None = None) -> float | None:
    """从 BIT* 的 cost_history [(t, cost), ...] 中找到首次 cost <= target 的时间.

    注意: 对于 2DOF 场景, cost_history 使用 OMPL 的欧氏距离,
    而 target_cost (SBF) 是 geodesic 距离 (torus 最短路径).
    geodesic <= Euclidean, 所以 cost_history 可能始终大于 target_cost,
    即使 BIT* 最终的 geodesic path_length 已经小于 target.

    Fallback: 如果 cost_history 未达到 target, 但 final_cost (geodesic)
    <= target_cost, 则 BIT* 确实达到了 SBF 质量, 返回 total_time 作为保守上界.

    Returns: 时间(秒), 如果 BIT* 从未达到则返回 None.
    """
    if not cost_history:
        # 无 cost_history, 但最终 geodesic cost 达标
        if final_cost is not None and total_time is not None and final_cost <= target_cost:
            return total_time
        return None
    for t, c in cost_history:
        if c <= target_cost:
            return t
    # Fallback: cost_history 未达到 (可能因为 Euclidean vs geodesic 度量差异),
    # 但最终 path_length (geodesic) <= target
    if final_cost is not None and total_time is not None and final_cost <= target_cost:
        return total_time
    return None  # never reached


def _safe_float(v, default=float('nan')):
    if v is None:
        return default
    try:
        f = float(v)
        return f if not math.isnan(f) else default
    except (TypeError, ValueError):
        return default


def _style_header(ws, row, ncols):
    """Style header row."""
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill(start_color="4472C4", end_color="4472C4",
                           fill_type="solid")
    hdr_align = Alignment(horizontal="center", vertical="center",
                          wrap_text=True)
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align


def _style_data(ws, start_row, end_row, ncols):
    """Alternate row colors for data rows."""
    even_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3",
                            fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="B4C6E7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for r in range(start_row, end_row + 1):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = center
            cell.border = border
            if (r - start_row) % 2 == 1:
                cell.fill = even_fill


def _style_avg_row(ws, row, ncols):
    """Bold + yellow background for average row."""
    avg_font = Font(bold=True, size=11)
    avg_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC",
                           fill_type="solid")
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = avg_font
        cell.fill = avg_fill


def create_2dof_sheet(wb: Workbook, results_path: Path):
    """Create the 2DOF comparison sheet."""
    with open(results_path, encoding="utf-8") as f:
        data = json.load(f)

    ws = wb.active
    ws.title = "2DOF (8-obs)"

    # Headers
    headers = [
        "Seed",
        "SBF-Dijkstra\nTime (ms)",
        "SBF-Dijkstra\nPath Length",
        "RRT-Connect\nTime (ms)",
        "RRT-Connect\nPath Length",
        "BIT*\nTotal Time (ms)",
        "BIT*\n1st Sol Time (ms)",
        "BIT*\nFinal Path Length",
        "BIT*\nTime to SBF Quality (ms)",
        "BIT* Reached\nSBF Quality?",
        "Note",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _style_header(ws, 1, len(headers))

    # Collect data per seed
    seeds = sorted(data["results"].keys(), key=int)
    row = 2
    sums = {k: [] for k in [
        "sbf_time", "sbf_len", "rrt_time", "rrt_len",
        "bit_time", "bit_1st", "bit_len", "bit_to_sbf"]}

    for seed_key in seeds:
        seed_results = data["results"][seed_key]
        sbf = next((r for r in seed_results if r["method"] == "SBF-Dijkstra"), None)
        rrt = next((r for r in seed_results if r["method"] == "OMPL-RRTConnect"), None)
        bit = next((r for r in seed_results if r["method"] == "OMPL-BITstar"), None)

        if not (sbf and sbf["success"] and rrt and bit and bit["success"]):
            continue

        sbf_time = sbf["time_ms"]
        sbf_len = sbf["path_length"]
        rrt_time = rrt["time_ms"] if rrt["success"] else None
        rrt_len = rrt["path_length"] if rrt["success"] else None
        bit_time = bit["time_ms"]
        bit_1st = bit["first_sol_ms"]
        bit_len = bit["path_length"]

        # BIT* cost_history -> time to reach SBF quality
        # 注意: cost_history 是 OMPL Euclidean 度量, sbf_len 是 geodesic 度量
        # 对于 2DOF torus 拓扑, geodesic < Euclidean, 需要 fallback 检查
        cost_history = bit.get("cost_history", [])
        t_reach = _find_time_to_reach_cost(
            cost_history, sbf_len,
            final_cost=bit_len,
            total_time=bit_time / 1000 if bit_time else None)  # bit_time is ms
        t_reach_ms = t_reach * 1000 if t_reach is not None else None
        reached = t_reach is not None

        # Determine if this was a fallback case (cost_history metric != geodesic)
        ch_reached = any(c <= sbf_len for _, c in cost_history) if cost_history else False
        is_fallback = reached and not ch_reached

        ws.cell(row=row, column=1, value=int(seed_key))
        ws.cell(row=row, column=2, value=round(sbf_time, 1))
        ws.cell(row=row, column=3, value=round(sbf_len, 3))
        ws.cell(row=row, column=4, value=round(rrt_time, 1) if rrt_time else "FAIL")
        ws.cell(row=row, column=5, value=round(rrt_len, 3) if rrt_len else "FAIL")
        ws.cell(row=row, column=6, value=round(bit_time, 1))
        ws.cell(row=row, column=7, value=round(bit_1st, 1))
        ws.cell(row=row, column=8, value=round(bit_len, 3))
        ws.cell(row=row, column=9, value=round(t_reach_ms, 1) if t_reach_ms else "N/A (>timeout)")
        ws.cell(row=row, column=10, value="Yes" if reached else "No")

        # Note: mark fallback cases
        note = ""
        if is_fallback:
            note = "geodesic < total time"
        ws.cell(row=row, column=11, value=note)

        # Color the reached cell
        if reached:
            ws.cell(row=row, column=10).font = Font(color="006100")
            ws.cell(row=row, column=10).fill = PatternFill(
                start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        else:
            ws.cell(row=row, column=10).font = Font(color="9C0006")
            ws.cell(row=row, column=10).fill = PatternFill(
                start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        sums["sbf_time"].append(sbf_time)
        sums["sbf_len"].append(sbf_len)
        if rrt_time is not None:
            sums["rrt_time"].append(rrt_time)
            sums["rrt_len"].append(rrt_len)
        sums["bit_time"].append(bit_time)
        sums["bit_1st"].append(bit_1st)
        sums["bit_len"].append(bit_len)
        if t_reach_ms is not None:
            sums["bit_to_sbf"].append(t_reach_ms)

        row += 1

    _style_data(ws, 2, row - 1, len(headers))

    # Average row
    ws.cell(row=row, column=1, value="AVG")
    def _avg(lst):
        return sum(lst) / len(lst) if lst else None

    for col, key, fmt in [
        (2, "sbf_time", 1), (3, "sbf_len", 3),
        (4, "rrt_time", 1), (5, "rrt_len", 3),
        (6, "bit_time", 1), (7, "bit_1st", 1),
        (8, "bit_len", 3),
    ]:
        v = _avg(sums[key])
        ws.cell(row=row, column=col, value=round(v, fmt) if v else "N/A")

    # bit_to_sbf average
    v = _avg(sums["bit_to_sbf"])
    n_reached = len(sums["bit_to_sbf"])
    n_total = len(sums["sbf_time"])
    ws.cell(row=row, column=9,
            value=f"{round(v, 1)} ({n_reached}/{n_total})" if v else "N/A")
    ws.cell(row=row, column=10,
            value=f"{n_reached}/{n_total}")

    _style_avg_row(ws, row, len(headers))

    # Column widths
    widths = [8, 16, 16, 16, 16, 18, 18, 18, 24, 16, 24]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "B2"


def create_panda_sheet(wb: Workbook, results_path: Path):
    """Create the Panda 10-obs comparison sheet.

    Includes SBF forest-build / query-only breakdown and reuse query times.
    """
    with open(results_path, encoding="utf-8") as f:
        data = json.load(f)

    ws = wb.create_sheet("Panda (10-obs)")

    headers = [
        "Seed",                           # A  1
        "SBF Forest\nBuild (s)",          # B  2
        "SBF Query\n(solve) (s)",         # C  3
        "SBF Total\nTime (s)",            # D  4
        "SBF\nPath Length",               # E  5
        "SBF Reuse\nQuery Avg (s)",       # F  6
        "RRT-Connect\nTime (s)",          # G  7
        "RRT-Connect\nPath Length",       # H  8
        "BIT*\nTotal Time (s)",           # I  9
        "BIT*\n1st Sol Time (s)",         # J  10
        "BIT*\n1st Sol Cost",             # K  11
        "BIT*\nFinal Path Length",        # L  12
        "BIT*\nTime to SBF Quality (s)",  # M  13
        "BIT* Reached\nSBF Quality?",     # N  14
        "Note",                           # O  15
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _style_header(ws, 1, len(headers))

    # Group standard results by seed
    seed_data = {}
    for r in data["results"]:
        seed = r["seed"]
        if seed not in seed_data:
            seed_data[seed] = {}
        seed_data[seed][r["planner"]] = r

    # Group reuse results by seed
    reuse_by_seed: dict[int, list] = {}
    for r in data.get("reuse_results", []):
        s = r["seed"]
        reuse_by_seed.setdefault(s, []).append(r)

    # Pre-collect RRT times for outlier detection
    all_rrt_times = []
    for seed in sorted(seed_data.keys()):
        sd = seed_data[seed]
        rrt = sd.get("OMPL-RRTConnect", {})
        if rrt.get("success"):
            t = rrt.get("planning_time", rrt.get("wall_clock", 0))
            all_rrt_times.append(t)
    rrt_median = float(np.median(all_rrt_times)) if all_rrt_times else None

    row = 2
    sums = {k: [] for k in [
        "sbf_grow", "sbf_solve", "sbf_time", "sbf_len", "sbf_reuse",
        "rrt_time", "rrt_len",
        "bit_time", "bit_1st_time", "bit_1st_cost", "bit_len", "bit_to_sbf"]}
    sums_rrt_no_outlier = {"rrt_time": [], "rrt_len": []}

    for seed in sorted(seed_data.keys()):
        sd = seed_data[seed]
        sbf = sd.get("SBF-Dijkstra", {})
        rrt = sd.get("OMPL-RRTConnect", {})
        bit = sd.get("OMPL-BITstar", {})

        if not sbf.get("success"):
            continue

        sbf_time = sbf.get("planning_time", sbf.get("wall_clock", 0))
        sbf_cost = _safe_float(sbf.get("cost"))

        # SBF timing breakdown
        pt = sbf.get("phase_times", {})
        sbf_grow = pt.get("grow", 0)
        sbf_solve = pt.get("solve", 0)

        # Reuse query times for this seed (non-first queries)
        reuse_qs = reuse_by_seed.get(seed, [])
        reuse_ok = [r for r in reuse_qs
                    if not r.get("is_first_query") and r.get("success")]
        reuse_avg = (sum(r["phase_times"]["solve"] for r in reuse_ok)
                     / len(reuse_ok)) if reuse_ok else None

        rrt_ok = rrt.get("success", False)
        rrt_time = rrt.get("planning_time", rrt.get("wall_clock", 0)) if rrt_ok else None
        rrt_cost = _safe_float(rrt.get("cost")) if rrt_ok else None

        rrt_outlier = False
        if rrt_ok and rrt_time is not None and rrt_median is not None:
            rrt_outlier = rrt_time > 3 * rrt_median

        bit_ok = bit.get("success", False)
        bit_time = bit.get("planning_time", bit.get("wall_clock", 0)) if bit_ok else None
        bit_cost = _safe_float(bit.get("cost")) if bit_ok else None
        bit_fst_time = _safe_float(bit.get("first_solution_time"))
        bit_fst_cost = _safe_float(bit.get("first_solution_cost"))
        cost_history = bit.get("cost_history", [])

        t_reach = _find_time_to_reach_cost(
            cost_history, sbf_cost,
            final_cost=bit_cost,
            total_time=bit_time) if bit_ok else None
        reached = t_reach is not None

        # ── Write row ──
        ws.cell(row=row, column=1, value=seed)
        ws.cell(row=row, column=2, value=round(sbf_grow, 3))
        ws.cell(row=row, column=3, value=round(sbf_solve, 3))
        ws.cell(row=row, column=4, value=round(sbf_time, 2))
        ws.cell(row=row, column=5, value=round(sbf_cost, 3))
        ws.cell(row=row, column=6,
                value=round(reuse_avg, 3) if reuse_avg is not None else "N/A")
        ws.cell(row=row, column=7,
                value=round(rrt_time, 2) if rrt_time else "FAIL")
        ws.cell(row=row, column=8,
                value=round(rrt_cost, 3) if rrt_cost and not math.isnan(rrt_cost) else "FAIL")
        ws.cell(row=row, column=9,
                value=round(bit_time, 2) if bit_time else "FAIL")
        ws.cell(row=row, column=10,
                value=round(bit_fst_time, 2) if not math.isnan(bit_fst_time) else "N/A")
        ws.cell(row=row, column=11,
                value=round(bit_fst_cost, 3) if not math.isnan(bit_fst_cost) else "N/A")
        ws.cell(row=row, column=12,
                value=round(bit_cost, 3) if bit_cost and not math.isnan(bit_cost) else "FAIL")

        note = ""
        if rrt_outlier:
            note = f"RRT outlier (>{3*rrt_median:.1f}s)"
            ws.cell(row=row, column=7).font = Font(color="9C0006", bold=True)

        ws.cell(row=row, column=13,
                value=f"{t_reach:.2f}" if reached else "N/A (>timeout)")
        ws.cell(row=row, column=14, value="Yes" if reached else "No")
        if reached:
            ws.cell(row=row, column=14).font = Font(color="006100")
            ws.cell(row=row, column=14).fill = PatternFill(
                start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        else:
            ws.cell(row=row, column=14).font = Font(color="9C0006")
            ws.cell(row=row, column=14).fill = PatternFill(
                start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        ws.cell(row=row, column=15, value=note)

        # Accumulators
        sums["sbf_grow"].append(sbf_grow)
        sums["sbf_solve"].append(sbf_solve)
        sums["sbf_time"].append(sbf_time)
        sums["sbf_len"].append(sbf_cost)
        if reuse_avg is not None:
            sums["sbf_reuse"].append(reuse_avg)
        if rrt_time is not None:
            sums["rrt_time"].append(rrt_time)
            sums["rrt_len"].append(rrt_cost)
            if not rrt_outlier:
                sums_rrt_no_outlier["rrt_time"].append(rrt_time)
                sums_rrt_no_outlier["rrt_len"].append(rrt_cost)
        if bit_time is not None:
            sums["bit_time"].append(bit_time)
            sums["bit_len"].append(bit_cost)
        if not math.isnan(bit_fst_time):
            sums["bit_1st_time"].append(bit_fst_time)
        if not math.isnan(bit_fst_cost):
            sums["bit_1st_cost"].append(bit_fst_cost)
        if t_reach is not None:
            sums["bit_to_sbf"].append(t_reach)

        row += 1

    _style_data(ws, 2, row - 1, len(headers))

    # ── Average row ──
    ws.cell(row=row, column=1, value="AVG")

    def _avg(lst):
        return sum(lst) / len(lst) if lst else None
    def _median(lst):
        if not lst:
            return None
        s = sorted(lst)
        n = len(s)
        return (s[n//2] + s[(n-1)//2]) / 2

    for col, key, fmt in [
        (2, "sbf_grow", 3), (3, "sbf_solve", 3),
        (4, "sbf_time", 2), (5, "sbf_len", 3),
        (6, "sbf_reuse", 3),
        (9, "bit_time", 2), (10, "bit_1st_time", 2),
        (11, "bit_1st_cost", 3), (12, "bit_len", 3),
    ]:
        v = _avg(sums[key])
        ws.cell(row=row, column=col, value=round(v, fmt) if v else "N/A")

    # RRT: avg (all) vs avg (excl outliers)
    n_outliers = len(sums["rrt_time"]) - len(sums_rrt_no_outlier["rrt_time"])
    if n_outliers > 0:
        v_no = _avg(sums_rrt_no_outlier["rrt_time"])
        ws.cell(row=row, column=7,
                value=f"{round(v_no, 2)} (excl {n_outliers} outlier)" if v_no else "N/A")
        v_len = _avg(sums_rrt_no_outlier["rrt_len"])
        ws.cell(row=row, column=8,
                value=round(v_len, 3) if v_len else "N/A")
    else:
        v = _avg(sums["rrt_time"])
        ws.cell(row=row, column=7, value=round(v, 2) if v else "N/A")
        v = _avg(sums["rrt_len"])
        ws.cell(row=row, column=8, value=round(v, 3) if v else "N/A")

    v = _avg(sums["bit_to_sbf"])
    n_reached = len(sums["bit_to_sbf"])
    n_total = len(sums["sbf_time"])
    ws.cell(row=row, column=13,
            value=f"{round(v, 2)} ({n_reached}/{n_total})" if v else "N/A")
    ws.cell(row=row, column=14,
            value=f"{n_reached}/{n_total}")

    _style_avg_row(ws, row, len(headers))

    widths = [8, 16, 16, 14, 14, 16, 16, 16, 16, 18, 18, 18, 24, 16, 28]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "B2"


def main():
    _ROOT = Path(__file__).resolve().parents[1]

    # Find latest 2DOF results with cost_history
    dof2_dirs = sorted(
        _ROOT.glob("experiments/output/sbf_vs_rrt_2dof_*/results.json"),
        key=lambda p: p.stat().st_mtime, reverse=True)
    dof2_path = dof2_dirs[0] if dof2_dirs else None

    # Panda 10-obs
    panda_path = _ROOT / "experiments" / "output" / "raw" / "panda_10obs_comparison.json"

    print(f"2DOF data:  {dof2_path}")
    print(f"Panda data: {panda_path}")

    wb = Workbook()

    if dof2_path and dof2_path.exists():
        create_2dof_sheet(wb, dof2_path)
    else:
        print("WARNING: No 2DOF results found")

    if panda_path.exists():
        create_panda_sheet(wb, panda_path)
    else:
        print("WARNING: No Panda 10-obs results found")

    out = _ROOT / "experiments" / "output" / "benchmark_comparison_v4.xlsx"
    wb.save(out)
    print(f"\nExcel saved: {out}")


if __name__ == "__main__":
    main()
